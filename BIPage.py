import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Font, Border, Side

class BIPage:
    def __init__(self, master):
        """
        初始化 BIPage 类的一个实例。

        Args:
            master (tk.Tk): 用作此界面的主窗口。
        """
        self.master = master
        self.master.title("生成福袋收益数据源")
        self.master.geometry("780x220")  # 确保窗口足够宽以显示所有内容

        # 创建一个Frame容器
        frame = tk.Frame(master, padx=10, pady=10)
        frame.grid(row=0, column=0, sticky="ew")

        # 福袋卡牌汇总的文件路径
        tk.Label(frame, text="福袋卡牌汇总的文件路径：").grid(
            row=0, column=0, sticky="w"
        )
        self.card_path_entry = tk.Entry(frame, width=60)
        self.card_path_entry.grid(row=0, column=1, padx=5, pady=5)
        tk.Button(frame, text="浏览", command=self.browse_card_path).grid(
            row=0, column=2, padx=5
        )

        # 福袋情况汇总的文件路径
        tk.Label(frame, text="福袋情况汇总的文件路径：").grid(
            row=1, column=0, sticky="w"
        )
        self.status_path_entry = tk.Entry(frame, width=60)
        self.status_path_entry.grid(row=1, column=1, padx=5, pady=5)
        tk.Button(frame, text="浏览", command=self.browse_status_path).grid(
            row=1, column=2, padx=5
        )

        # 上一期Page1数据源的文件路径
        tk.Label(frame, text="上一期线上福袋收益-Page1数据源的文件路径：").grid(
            row=2, column=0, sticky="w"
        )
        self.previous_path_entry = tk.Entry(frame, width=60)
        self.previous_path_entry.grid(row=2, column=1, padx=5, pady=5)
        tk.Button(frame, text="浏览", command=self.browse_previous_path).grid(
            row=2, column=2, padx=5
        )

        # 输出路径
        tk.Label(frame, text="线上福袋收益-Page1数据源（.xlsx）的输出路径：").grid(
            row=3, column=0, sticky="w"
        )
        self.output_path_entry = tk.Entry(frame, width=60)
        self.output_path_entry.grid(row=3, column=1, padx=5, pady=5)
        tk.Button(frame, text="浏览", command=self.browse_output_path).grid(
            row=3, column=2, padx=5
        )

        # 归属主体，添加占位符
        tk.Label(frame, text="归属主体：").grid(row=4, column=0, sticky="w")
        self.body_entry = tk.Entry(frame, width=60, fg="grey")
        self.body_entry.grid(row=4, column=1, padx=5, pady=5)
        self.body_entry.insert(
            0,
            "生成总体的福袋收益，则输入 总体 ，否则输入所需的归属主体（白鸽、骨头...）",
        )
        self.body_entry.bind("<FocusIn>", self.on_entry_click)
        self.body_entry.bind("<FocusOut>", self.on_focusout)

        # 生成按钮
        tk.Button(frame, text="生成Page1数据源", command=self.generate_report).grid(
            row=5, column=1, pady=10
        )

    def on_entry_click(self, event):
        """
        处理输入框的点击事件，清除默认文本。

        Args:
            event: 事件对象。
        """
        if (
            self.body_entry.get()
            == "生成总体的福袋收益，则输入 总体 ，否则输入所需的归属主体（白鸽、骨头...）"
        ):
            self.body_entry.delete(0, "end")  # delete all the text in the entry
            self.body_entry.insert(0, "")  # Insert blank for user input
            self.body_entry.config(fg="black")

    def on_focusout(self, event):
        """
        处理输入框失去焦点的事件，如果为空则重新显示默认文本。

        Args:
            event: 事件对象。
        """
        if self.body_entry.get() == "":
            self.body_entry.insert(
                0,
                "生成总体的福袋收益，则输入 总体 ，否则输入所需的归属主体（白鸽、骨头...）",
            )
            self.body_entry.config(fg="grey")

    def browse_card_path(self):
        """
        打开文件对话框，让用户选择福袋卡牌汇总的文件路径。
        """
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.card_path_entry.delete(0, tk.END)
            self.card_path_entry.insert(0, path)

    def browse_status_path(self):
        """
        打开文件对话框，让用户选择福袋卡牌汇总的文件路径。
        """
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.status_path_entry.delete(0, tk.END)
            self.status_path_entry.insert(0, path)

    def browse_previous_path(self):
        """
        打开文件对话框，让用户选择上一期线上福袋收益-Page1数据源的文件路径。
        """
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.previous_path_entry.delete(0, tk.END)
            self.previous_path_entry.insert(0, path)

    def browse_output_path(self):
        """
        打开文件对话框，让用户选择输出文件的路径，并设置默认扩展名为.xlsx。
        """
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            self.output_path_entry.delete(0, tk.END)
            self.output_path_entry.insert(0, path)

    def generate_report(self):
        """
        根据用户输入的路径和参数生成Page1数据源，并显示成功或失败的消息。
        """
        try:
            card_file_path = self.card_path_entry.get()
            status_file_path = self.status_path_entry.get()
            previous_file_path = self.previous_path_entry.get()
            output_path = self.output_path_entry.get()
            body = self.body_entry.get()
            self.BIpage1(
                card_file_path, status_file_path, previous_file_path, output_path, body
            )
            messagebox.showinfo("成功", "成功生成Page1数据源")
        except Exception as e:
            messagebox.showerror("生成失败", f"生成Page1数据源失败：{str(e)}")

    def BIpage1(
        self,
        card_file_path: str | Path,
        status_file_path: str | Path,
        previous_file_path: str | Path,
        output_path: str | Path,
        body: str = "总体",
    ):
        """
        生成BIPage1数据源


        Args:
            card_file_path (str | Path): 福袋卡牌汇总的文件路径
            status_file_path (str | Path): 福袋情况汇总的文件路径
            previous_file_path (str | Path): 上一期线上福袋收益-Page1数据源的文件路径
            output_path (str | Path): 线上福袋收益-Page1数据源的文件输出路径
            body (str): 报告的主体，用于指定报告的具体内容或范围，默认值为"总体"

        Returns:

        """

        card_excel = pd.read_excel(card_file_path, sheet_name=0)
        status_excel = pd.read_excel(status_file_path, sheet_name=0)
        previous_excel = pd.read_excel(previous_file_path, sheet_name="项目总体营收")
        previous_derived = pd.read_excel(previous_file_path, sheet_name="derived")

        previous_data = previous_excel

        # 提取出对应的归属主体中的数据
        if body == "总体":
            card_data = card_excel
            status_data = status_excel

        else:
            card_data = card_excel[card_excel["归属主体"] == body]
            status_data = status_excel[status_excel["归属主体"] == body]

        ##### Sheet1 -- 福袋汇总
        sheet1 = status_data.loc[status_data["卡牌品类"] == "PTCG"].copy()
        sheet1["上架日期"] = pd.to_datetime(sheet1["上架日期"], format="%Y/%m/%d")
        sheet1["上架日期"] = sheet1["上架日期"].dt.strftime("%Y/%m/%d")

        ##### Sheet2 -- 项目总体营收
        status_data.loc[:, "上架日期"] = pd.to_datetime(status_data["上架日期"], format="%Y/%m/%d")
        start = status_data["上架日期"].min()
        end = datetime.now()

        # 起始日 - start
        # 结算日 - end
        # 累计购卡支出 - total_card_purchase_expense
        # 累计玩法奖品支出 - total_game_prize_expense
        # 未支付卡片成本 - unpaid_card_cost
        # 库存货物成本 - inventory_cost
        # 正投放货物成本 - active_deployment_cost
        # 累计实际消耗成本 - total_actual_consumption_cost
        # 数据周期长度 - data_period_length
        # 累计营销额 - total_marketing_amount
        # 上一期累计营销额 - previous_total_marketing_amount
        # 累计销售盈利 - total_sales_profit
        # 上一期累计销售盈利 - previous_total_sales_profit
        # 累计内部垫抽金额 - total_internal_advance_amount
        # 累计流程成本（平台及税收） - total_process_cost
        # 累计福袋净收益（除运营成本外） - total_lucky_bag_net_profit
        # 上一期累计福袋净收益（除运营成本外） - previous_total_lucky_bag_net_profit
        # 收益率（消耗成本） - roi_consumption_cost
        # 上一期收益率（消耗成本） - previous_roi_consumption_cost

        # 挂靠主体
        GKBody = body

        # 计算起始日
        start_date = start.strftime("%Y/%m/%d")

        # 计算结算日
        end_date = end.strftime("%Y/%m/%d")

        # 累计购卡支出
        total_card_purchase_expense = ""

        # 累计玩法奖品支出
        total_game_prize_expense = ""

        # 计算未支付卡片成本
        filtered_data = card_data[
            (card_data["状态"].isin(["已过审", "已入袋"]))
            & (card_data["备注"].str.contains("未支付成本"))
        ]
        unpaid_card_cost = filtered_data["成本价格"].sum()

        # 计算库存货物成本
        filtered_data = card_data[(card_data["状态"].isin(["已过审"]))]
        inventory_cost = filtered_data["成本价格"].sum()

        # 计算正投放货物成本
        filtered_data = card_data[(card_data["状态"].isin(["已入袋"]))]
        active_deployment_cost = filtered_data["成本价格"].sum()

        # 计算累计实际消耗成本
        filtered_data = card_data[(card_data["状态"].isin(["已抽走"]))]
        total_actual_consumption_cost = filtered_data["成本价格"].sum()

        # 计算数据周期长度
        data_period_length = (end - start).days

        # 计算累计营销额
        total_marketing_amount = sheet1["销售额"].sum()

        # 计算上一期累计营销额
        previous_total_marketing_amount = previous_data["累计营销额"].iloc[0]

        # 计算累计销售盈利
        total_sales_profit = total_marketing_amount - total_actual_consumption_cost

        # 计算上一期累计销售盈利
        previous_total_sales_profit = previous_data["累计销售盈利"].iloc[0]

        # 计算累计内部垫抽金额
        total_internal_advance_amount = sheet1["垫抽金额"].sum()

        # 计算累计流程成本（平台及税收）
        total_process_cost = (
            total_marketing_amount * 0.05 + total_marketing_amount * 0.95 / 1.01 * 0.01
        ).round(2)

        # 计算累计福袋净收益（除运营成本外）
        total_lucky_bag_net_profit = (
            total_sales_profit - total_internal_advance_amount - total_process_cost
        )

        # 上一期累计福袋净收益（除运营成本外）
        previous_total_lucky_bag_net_profit = previous_data[
            "累计福袋净收益（除运营成本外）"
        ].iloc[0]

        # 计算收益率（消耗成本）
        roi_consumption_cost = (
            total_lucky_bag_net_profit / total_actual_consumption_cost
        )

        # 计算上一期收益率（消耗成本）
        previous_roi_consumption_cost = previous_data["收益率（消耗成本）"].iloc[0]

        sheet2 = pd.DataFrame(
            {
                "挂靠主体": [GKBody],
                "计算起始日": [start_date],
                "计算结算日": [end_date],
                "累计购卡支出": [total_card_purchase_expense],  # 这里需要填入实际计算值
                "累计玩法奖品支出": [
                    total_game_prize_expense
                ],  # 这里需要填入实际计算值
                "未支付卡片成本": [unpaid_card_cost],
                "库存货物成本": [inventory_cost],
                "正投放货物成本": [active_deployment_cost],
                "累计实际消耗成本": [total_actual_consumption_cost],
                "数据周期长度": [data_period_length],
                "累计营销额": [total_marketing_amount],
                "上一期累计营销额": [
                    previous_total_marketing_amount
                ],  # 需要填入实际值或留空
                "累计销售盈利": [total_sales_profit],
                "上一期累计销售盈利": [
                    previous_total_sales_profit
                ],  # 需要填入实际值或留空
                "累计内部垫抽金额": [total_internal_advance_amount],
                "累计流程成本（平台及税收）": [total_process_cost],
                "累计福袋净收益（除运营成本外）": [total_lucky_bag_net_profit],
                "上一期累计福袋净收益（除运营成本外）": [
                    previous_total_lucky_bag_net_profit
                ],  # 需要填入实际值或留空
                "收益率（消耗成本）": [roi_consumption_cost],
                "上一期收益率（消耗成本）": [
                    previous_roi_consumption_cost
                ],  # 需要填入实际值或留空
            }
        )

        ##### Sheet3 -- 垫抽情况统计
        sheet3 = status_data[["福袋名", "垫抽数", "净收益率", "净收益金额"]]
        sheet3 = sheet3.rename(columns={"净收益率": "源数据1", "净收益金额": "源数据2"})
        # 添加新列并进行数值格式化
        sheet3["收益率"] = sheet3["源数据1"].round(4)  # 第5列，保留4位小数
        sheet3["营收/亏损"] = sheet3["源数据2"].round(2)  # 第6列，保留2位小数

        ##### Sheet4 -- 货物价值盈亏统计
        cumulative_cost = card_data["成本价格"].sum()
        cumulative_market_value = card_data["最新市价"].sum()
        price_change_profit_loss_ratio = (cumulative_market_value / cumulative_cost) - 1
        sheet4 = pd.DataFrame(
            {
                "货物累计成本价": [cumulative_cost],
                "货物累计市价": [cumulative_market_value],
                "货物价格变动盈亏比": [price_change_profit_loss_ratio],
            }
        )

        ##### Sheet5 -- 提现情况统计
        history_cumulative_sales = status_data["销售额"].sum()
        history_cumulative_withdrawable_amount = (
            history_cumulative_sales * 0.95
        ).round(2)
        sheet5 = pd.DataFrame(
            {
                "历史累计销售额": [history_cumulative_sales],
                "历史累计可提现金额": [history_cumulative_withdrawable_amount],
                "历史已提现金额": "",
                "可提现金额": "",
            }
        )

        ##### Sheet6 -- 周收益统计
        status_data.loc[:, "上架日期"] = pd.to_datetime(status_data["上架日期"], format="%Y/%m/%d")
        status_data.set_index("上架日期", inplace=True)

        # 计算每周的收益金额总和
        week_profit_sum = (
            status_data["净收益金额"].resample("W-SUN").sum().reset_index()
        )
        # 计算每周的垫抽金额总和
        advance_draw_sum = status_data["垫抽金额"].resample("W-SUN").sum().reset_index()

        # 创建 sheet6 DataFrame
        sheet6 = pd.DataFrame(
            {
                "周数": range(1, len(week_profit_sum) + 1),
                "当周收益金额": week_profit_sum["净收益金额"],
            }
        )

        # 计算累计收益金额
        sheet6["累计收益金额"] = sheet6["当周收益金额"] + sheet6[
            "当周收益金额"
        ].cumsum().shift(1).fillna(0)
        # 计算当周收益金额含垫抽
        sheet6["当周收益金额含垫抽"] = (
            week_profit_sum["净收益金额"] + advance_draw_sum["垫抽金额"]
        )
        # 累计收益金额含垫抽
        sheet6["累计收益金额含垫抽"] = sheet6["当周收益金额含垫抽"] + sheet6[
            "当周收益金额含垫抽"
        ].cumsum().shift(1).fillna(0)

        # 创建 sheet7 DataFrame



        with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
            sheet1.to_excel(writer, sheet_name="福袋汇总", index=False)
            sheet2.to_excel(writer, sheet_name="项目总体营收", index=False)
            sheet3.to_excel(writer, sheet_name="垫抽情况统计", index=False)
            sheet4.to_excel(writer, sheet_name="货物价值盈亏统计", index=False)
            sheet5.to_excel(writer, sheet_name="提现情况统计", index=False)
            sheet6.to_excel(writer, sheet_name="周收益统计", index=False)

        # 打开现有的 Excel 文件
        wb = openpyxl.load_workbook(output_path)

        # 删除旧的 sheet3
        if "derived" in wb.sheetnames:
            del wb["derived"]

        # 创建新的 derived 页
        derived_sheet = wb.create_sheet(title="derived", index=2)

        # 定义标题和格式
        headers = [
            "公司投入资金", "起始日", "结算日", "累计运营成本", "累计净收益",
            "收益率（总成本）", "上一期收益率（总成本）", "收益率（总成本）月化",
            "上一期收益率（总成本）月化", "收益率（总成本）年化", "上一期收益率（总成本）年化", "数据周期长度"
        ]

        # 设置字体和对齐方式
        header_font = Font(name='微软雅黑', size=11, bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # 将标题写入A1到L1并设置样式
        for col_num, header in enumerate(headers, 1):
            cell = derived_sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            # 设置列宽
            derived_sheet.column_dimensions[cell.column_letter].width = len(header) + 2  # 适当加2个字符宽度

        # 填写公式和数据
        derived_sheet['B2'] = '=项目总体营收!B2'
        derived_sheet['C2'] = '=项目总体营收!C2'
        derived_sheet['B2'].number_format = 'YYYY-MM-DD'
        derived_sheet['C2'].number_format = 'YYYY-MM-DD'
        derived_sheet['E2'] = '=项目总体营收!R2 - derived!D2'
        derived_sheet['F2'] = '=ROUND(E2/A2, 4)'
        derived_sheet['H2'] = '=ROUND(F2/(C2-B2)*30,4)'
        derived_sheet['J2'] = '=H2*12'
        derived_sheet['L2'] = '=C2-B2'

        derived_sheet['G2'] = previous_derived["收益率（总成本）"].iloc[0]
        derived_sheet['I2'] = previous_derived["收益率（总成本）月化"].iloc[0]
        derived_sheet['K2'] = previous_derived["收益率（总成本）年化"].iloc[0]

        derived_sheet['F2'].number_format = '0.00%'
        derived_sheet['G2'].number_format = '0.00%'
        derived_sheet['H2'].number_format = '0.00%'
        derived_sheet['I2'].number_format = '0.00%'
        derived_sheet['J2'].number_format = '0.00%'
        derived_sheet['K2'].number_format = '0.00%'

        # 设置列宽
        for col in derived_sheet.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 12
            derived_sheet.column_dimensions[column].width = adjusted_width

        # 设置第二行的字体
        row_font = Font(name='微软雅黑', size=11, bold=True)
        for col_num in range(1, len(headers) + 1):
            cell = derived_sheet.cell(row=2, column=col_num)
            cell.font = row_font
            cell.alignment = header_alignment


        # 保存修改后的工作簿
        wb.save(output_path)



# 创建窗口并运行应用
root = tk.Tk()
app = BIPage(root)
root.mainloop()
